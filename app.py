
import uuid
from datetime import datetime
from pathlib import Path

import streamlit as st
from openpyxl import Workbook, load_workbook


# ==============================================================================
# APP CONFIGURATION
# ==============================================================================

st.set_page_config(
    page_title="Bangla Pragmatics Data Collection",
    page_icon="🗣️",
    layout="wide",
)


# ==============================================================================
# FILE CONFIGURATION
# ==============================================================================

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

EXCEL_FILE = DATA_DIR / "bangla_pragmatics_responses.xlsx"


# ==============================================================================
# SCENARIOS
# ==============================================================================

SCENARIOS = [

    (
        "Direct Request",
        (
            "আপনি আপনার ঘরে পড়াশোনা করছেন। পাশের ঘরে আপনার ছোট ভাই/বোন খুব"
            " জোরে গান বাজাচ্ছে এবং আপনি মনোযোগ দিতে পারছেন না।"
        ),
    ),

    (
        "Indirect Request",
        (
            "আপনি একজন বন্ধুর সঙ্গে একটি ঘরে বসে আছেন। জানালা খোলা এবং বাইরে"
            " অনেক ঠান্ডা।"
        ),
    ),

    (
        "Asking Someone to Move",
        (
            "আপনি বাসে বসতে যাচ্ছেন। পাশের আসনে একজন তার ব্যাগ রেখেছে এবং আপনি"
            " সেখানে বসতে চান।"
        ),
    ),

    (
        "Request for Help",
        "আপনার হাতে কয়েকটি ভারী ব্যাগ এবং আপনার বন্ধু কাছেই দাঁড়িয়ে আছে।",
    ),

    (
        "Indirect Request to Stop",
        (
            "গুরুত্বপূর্ণ পরীক্ষার প্রস্তুতির সময় আপনার বন্ধু বারবার আপনাকে"
            " মেসেজ পাঠাচ্ছে।"
        ),
    ),

    (
        "Direct Refusal",
        (
            "আপনার বন্ধু আপনাকে একটি পার্টিতে যেতে বলেছে, কিন্তু পরদিন সকালে"
            " আপনার গুরুত্বপূর্ণ পরীক্ষা।"
        ),
    ),

    (
        "Indirect Refusal",
        (
            "একজন বন্ধু জিজ্ঞেস করল: “কাল আমার সাথে ঘুরতে যাবে?” কিন্তু আপনার"
            " আগে থেকেই অন্য একটি কাজ আছে।"
        ),
    ),

    (
        "Polite Refusal",
        (
            "একজন সহকর্মী আপনাকে আজ অতিরিক্ত কাজ করতে বলেছে, কিন্তু আপনি"
            " ইতিমধ্যে অনেক কাজের চাপে আছেন।"
        ),
    ),

    (
        "Avoiding an Invitation",
        (
            "আপনাকে এমন একজন ব্যক্তি ডিনারে আমন্ত্রণ জানিয়েছে যাকে আপনি খুব বেশি"
            " চেনেন না, কিন্তু আপনি যেতে চান না।"
        ),
    ),

    (
        "Refusing to Lend",
        (
            "একজন বন্ধু আপনার ল্যাপটপ ধার চাইছে, কিন্তু আপনি ল্যাপটপটি দিতে চান"
            " না।"
        ),
    ),

    (
        "Sarcasm: Late Arrival",
        (
            "আপনার বন্ধু বিকেল ৫টায় আসার কথা বলেছিল, কিন্তু সে রাত ৭টায় এসে"
            " পৌঁছাল।"
        ),
    ),

    (
        "Sarcasm: Poor Performance",
        (
            "আপনার বন্ধু একটি কাজকে খুব সহজ বলেছিল, কিন্তু শেষ পর্যন্ত সে কাজটি"
            " সম্পূর্ণ করতে ব্যর্থ হয়েছে।"
        ),
    ),

    (
        "Sarcasm: Messy Room",
        (
            "আপনার ভাই/বোন ঘর পরিষ্কার করবে বলেছিল, কিন্তু ঘরটি আরও বেশি এলোমেলো"
            " করে ফেলেছে।"
        ),
    ),

    (
        "Sarcasm: Broken Promise",
        (
            "আপনার বন্ধু কয়েকবার আপনাকে ফোন করবে বলেছিল, কিন্তু কখনো ফোন করেনি।"
            " অবশেষে সে ফোন করল।"
        ),
    ),

    (
        "Sarcasm: Obvious Mistake",
        (
            "আপনার বন্ধু একটি খুব স্পষ্ট ভুল করেছে এবং বলছে, “আমি কখনো ভুল করি"
            " না।”"
        ),
    ),

    (
        "Irony: Rain",
        (
            "আপনি সারাদিনের জন্য একটি বাইরের অনুষ্ঠান আয়োজন করেছেন, কিন্তু"
            " অনুষ্ঠান শুরুর সময় প্রচণ্ড বৃষ্টি শুরু হলো।"
        ),
    ),

    (
        "Irony: Technology Failure",
        (
            "আপনি নতুন একটি ফোন কিনেছেন কারণ আপনি আশা করেছিলেন এটি খুব ভালো কাজ"
            " করবে, কিন্তু প্রথম দিনেই ফোনটি নষ্ট হয়ে গেল।"
        ),
    ),

    (
        "Irony: Perfect Timing",
        "আপনি বাসস্ট্যান্ডে পৌঁছানোর ঠিক পরেই আপনার বাসটি ছেড়ে দিল।",
    ),

    (
        "Irony: Group Project",
        (
            "আপনার গ্রুপের সদস্যরা প্রায় কোনো কাজ না করে শেষ মুহূর্তে ডেডলাইনের"
            " এক মিনিট আগে অ্যাসাইনমেন্ট জমা দিল।"
        ),
    ),

    (
        "Irony: Another Problem",
        (
            "আপনি অনেক কষ্টে একটি কঠিন কাজ শেষ করলেন, কিন্তু সঙ্গে সঙ্গেই আরেকটি"
            " বড় সমস্যা তৈরি হলো।"
        ),
    ),

    (
        "Rhetorical Question: Repeated Mistake",
        (
            "আপনার বন্ধু একই ভুল বারবার করছে, যদিও আপনি তাকে বিষয়টি অনেকবার"
            " বুঝিয়েছেন।"
        ),
    ),

    (
        "Rhetorical Question: Exam",
        (
            "একজন ব্যক্তি পরীক্ষার আগের দিন পড়াশোনা না করে অভিযোগ করছে যে তাকে"
            " কেন পড়তে হবে।"
        ),
    ),

    (
        "Rhetorical Question: Broken Object",
        (
            "আপনার ভাই/বোন কিছু ভেঙে ফেলেছে এবং এখন এমন আচরণ করছে যেন সে কিছুই"
            " জানে না।"
        ),
    ),

    (
        "Rhetorical Question: Tiredness",
        (
            "আপনি সারাদিন কাজ করার পর ক্লান্ত, কিন্তু আপনার বন্ধু বারবার জিজ্ঞেস"
            " করছে কেন আপনি ক্লান্ত।"
        ),
    ),

    (
        "Rhetorical Question: Ignored Advice",
        (
            "আপনি কাউকে বারবার সতর্ক করেছিলেন, সে আপনার কথা শোনেনি এবং এখন ঠিক"
            " সেই সমস্যায় পড়েছে।"
        ),
    ),

    (
        "Genuine Praise: Project",
        "আপনার বন্ধু একটি কঠিন প্রজেক্ট সফলভাবে সম্পন্ন করেছে।",
    ),

    (
        "Genuine Praise: Academic Result",
        "আপনার ছোট ভাই/বোন পরীক্ষায় খুব ভালো ফল করেছে।",
    ),

    (
        "Appreciation: Help",
        "একজন সহকর্মী আপনাকে একটি গুরুত্বপূর্ণ কাজ শেষ করতে সাহায্য করেছে।",
    ),

    (
        "Praise: Appearance",
        "আপনার বন্ধু নতুন একটি পোশাক পরেছে এবং তাকে খুব সুন্দর দেখাচ্ছে।",
    ),

    (
        "Praise: Presentation",
        "আপনি আপনার বন্ধুর একটি খুব ভালো প্রেজেন্টেশন শুনলেন।",
    ),

    (
        "Complaint: Noise",
        "আপনার প্রতিবেশী সপ্তাহে কয়েকবার গভীর রাত পর্যন্ত জোরে গান বাজায়।",
    ),

    (
        "Complaint: Late Delivery",
        (
            "আপনি অনলাইনে কিছু অর্ডার করেছিলেন, কিন্তু প্রতিশ্রুত সময়ের কয়েক দিন"
            " পরে সেটি এসেছে।"
        ),
    ),

    (
        "Complaint: Group Member",
        (
            "আপনার গ্রুপের একজন সদস্য বারবার তার দায়িত্বের কাজ শেষ করছে না।"
        ),
    ),

    (
        "Complaint: Restaurant",
        (
            "আপনি রেস্টুরেন্টে খাবার অর্ডার করেছেন, কিন্তু খাবারটি আপনার প্রত্যাশার"
            " তুলনায় খুব খারাপ।"
        ),
    ),

    (
        "Complaint: Cancelled Plans",
        "আপনার বন্ধু বারবার শেষ মুহূর্তে আপনাদের পরিকল্পনা বাতিল করছে।",
    ),

    (
        "Ambiguous: Good News",
        "আপনার বন্ধু সবেমাত্র একটি খুব ভালো খবর পেয়েছে।",
    ),

    (
        "Ambiguous: Serious Mistake",
        "আপনার বন্ধু সবেমাত্র একটি গুরুতর ভুল করেছে।",
    ),

    (
        "Ambiguous: Uncertain Help",
        (
            "কেউ আপনাকে সাহায্য করবে বলেছে, কিন্তু সে সত্যিই সাহায্য করবে কি না তা"
            " নিয়ে আপনি নিশ্চিত নন।"
        ),
    ),

    (
        "Ambiguous: Waiting",
        (
            "আপনি এমন একজনের জন্য অপেক্ষা করছেন যে ইতিমধ্যে ৩০ মিনিট দেরি করেছে এবং"
            " সে এখন এসে পৌঁছেছে।"
        ),
    ),

    (
        "Ambiguous: Indirect No",
        (
            "কেউ আপনাকে এমন কিছু করতে বলেছে যা আপনি করতে চান না, কিন্তু আপনি"
            " সরাসরি “না” বলতে চান না।"
        ),
    ),

    (
        "Doubt",
        (
            "আপনার বন্ধু বলছে সে মাত্র ৩০ মিনিটে একটি অনেক বড় অ্যাসাইনমেন্ট শেষ"
            " করেছে।"
        ),
    ),

    (
        "Surprise",
        "বহু বছর পর আপনি হঠাৎ একজন পুরোনো বন্ধুর সঙ্গে দেখা করলেন।",
    ),

    (
        "Disappointment",
        (
            "একটি গুরুত্বপূর্ণ অনুষ্ঠানে আপনি কারও সহযোগিতা আশা করেছিলেন, কিন্তু সে"
            " আসেনি।"
        ),
    ),

    (
        "Sympathy",
        "আপনার বন্ধু জানাল যে সে একটি গুরুত্বপূর্ণ পরীক্ষায় ফেল করেছে।",
    ),

    (
        "Frustration",
        (
            "আপনি একই সমস্যা কয়েকবার বুঝিয়েছেন, কিন্তু অন্য ব্যক্তি এখনও"
            " বিষয়টি বুঝতে পারছে না।"
        ),
    ),

    (
        "Teacher–Student",
        (
            "আপনি একটি অ্যাসাইনমেন্ট দেরিতে জমা দিয়েছেন এবং শিক্ষক জিজ্ঞেস করলেন"
            " কেন দেরি হয়েছে।"
        ),
    ),

    (
        "Workplace",
        (
            "আপনার সুপারভাইজার আপনাকে আরও একটি কাজ দিলেন, অথচ আপনার হাতে ইতিমধ্যে"
            " কয়েকটি জরুরি কাজ আছে।"
        ),
    ),

    (
        "Family",
        (
            "আপনার পরিবার আপনাকে একটি অনুষ্ঠানে যেতে বলছে, কিন্তু আপনার"
            " গুরুত্বপূর্ণ একাডেমিক কাজ আছে।"
        ),
    ),

    (
        "Stranger",
        (
            "ভিড়ের মধ্যে একজন অপরিচিত ব্যক্তি আপনাকে ধাক্কা দিলেন এবং ক্ষমা"
            " চাইলেন না।"
        ),
    ),

    (
        "Keeping a Secret",
        (
            "আপনার বন্ধু আপনাকে একটি ব্যক্তিগত বিষয় বলেছে এবং কাউকে না বলতে"
            " অনুরোধ করেছে। পরে অন্য একজন বন্ধু বিষয়টি সম্পর্কে জানতে চাইল।"
        ),
    ),
]


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def safe_text(x):
    """Remove invalid characters and convert to clean text."""
    return str(x).replace("\x00", "").strip()


def participant_id_from_state(participant):
    """Return participant_id only if valid participant data exists."""
    if not isinstance(participant, dict):
        return None
    return participant.get("participant_id")


# ==============================================================================
# INITIALIZE EXCEL FILE
# ==============================================================================

def init_excel():

    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()

    # --------------------------------------------------------------------------
    # Sheet 1: Participant Responses
    # --------------------------------------------------------------------------

    if "Responses" in wb.sheetnames:
        ws = wb["Responses"]
    else:
        ws = wb.active
        ws.title = "Responses"

    if ws.max_row == 1 and ws["A1"].value is None:
        response_headers = [

            "Response_ID",
            "Participant_ID",
            "Timestamp",

            "Participant_Age_Group",
            "Education",
            "Bangla_Usage",
            "Bangla_Variety",

            "Scenario_ID",
            "Scenario_Category",
            "Context",

            "Utterance",
            "Intended_Meaning",
            "Primary_Intent_Category",

            "Scenario_Response_Time_Seconds",

        ]
        ws.append(response_headers)

    # --------------------------------------------------------------------------
    # Sheet 2: HCI / Usability Feedback
    # --------------------------------------------------------------------------

    if "Usability_Feedback" not in wb.sheetnames:
        usability_ws = wb.create_sheet("Usability_Feedback")
    else:
        usability_ws = wb["Usability_Feedback"]

    if usability_ws.max_row == 1 and usability_ws["A1"].value is None:
        usability_headers = [

            "Participant_ID",

            "Completion_Time_Minutes",

            "U1_Instructions_Clear",
            "U2_Scenarios_Understandable",
            "U3_Interface_Easy",
            "U4_Example_Helpful",
            "U5_Progress_Useful",
            "U6_Easy_to_Express_Meaning",
            "U7_Questionnaire_Length_Reasonable",
            "U8_Comfortable_Using_System",
            "U9_Willing_to_Reuse",
            "U10_Overall_Satisfaction",

            "Open_Feedback",

        ]
        usability_ws.append(usability_headers)

    wb.save(EXCEL_FILE)


# ==============================================================================
# SAVE SCENARIO RESPONSE
# ==============================================================================

def save_response(

    participant,

    scenario_id,
    category,
    context,

    utterance,
    intended,
    intent_cat,

    response_time,

):

    participant_id = participant_id_from_state(participant)

    if participant_id is None:
        st.warning(
            "Participant information is missing. Please restart the questionnaire and begin again."
        )
        return False

    init_excel()

    wb = load_workbook(EXCEL_FILE)

    ws = wb["Responses"]

    ws.append([

        str(uuid.uuid4()),

        participant_id,

        datetime.now().isoformat(timespec="seconds"),

        participant["age"],
        participant["education"],
        participant["usage"],
        participant["variety"],

        scenario_id,
        category,
        context,

        safe_text(utterance),
        safe_text(intended),
        safe_text(intent_cat),

        round(response_time, 2),

    ])

    wb.save(EXCEL_FILE)
    return True


# ==============================================================================
# SAVE HCI USABILITY FEEDBACK
# ==============================================================================

def save_usability_feedback(

    participant_id,

    completion_time,

    u1,
    u2,
    u3,
    u4,
    u5,
    u6,
    u7,
    u8,
    u9,
    u10,

    feedback,

):

    init_excel()

    wb = load_workbook(EXCEL_FILE)

    ws = wb["Usability_Feedback"]

    ws.append([

        participant_id,

        round(completion_time, 2),

        u1,
        u2,
        u3,
        u4,
        u5,
        u6,
        u7,
        u8,
        u9,
        u10,

        safe_text(feedback),

    ])

    wb.save(EXCEL_FILE)


# ==============================================================================
# CREATE EXCEL FILE
# ==============================================================================

init_excel()


# ==============================================================================
# SESSION STATE INITIALIZATION
# ==============================================================================

if "started" not in st.session_state:
    st.session_state.started = False

if "participant" not in st.session_state:
    st.session_state.participant = None

if "index" not in st.session_state:
    st.session_state.index = 0

if "answers" not in st.session_state:
    st.session_state.answers = {}

if "scenario_start_time" not in st.session_state:
    st.session_state.scenario_start_time = None

if "usability_completed" not in st.session_state:
    st.session_state.usability_completed = False


# ==============================================================================
# MAIN TITLE
# ==============================================================================

st.title("🗣️ Bangla Pragmatics Data Collection")

st.caption(
    "Context + Natural Utterance + Intended Meaning"
)


# ==============================================================================
# PARTICIPANT INFORMATION PAGE
# ==============================================================================

if not st.session_state.started:

    st.markdown(
        """
        ## গবেষণা সম্পর্কিত তথ্য

        এই গবেষণায় আপনাকে কিছু দৈনন্দিন বাংলা পরিস্থিতি দেওয়া হবে।

        প্রতিটি পরিস্থিতির জন্য আপনাকে লিখতে হবে:

        **১. আপনি স্বাভাবিকভাবে কী বলতেন।**

        **২. এই কথা বলে আপনি আসলে কী বোঝাতে চেয়েছেন।**

        এখানে কোনো সঠিক বা ভুল উত্তর নেই।

        অনুগ্রহ করে স্বাভাবিক দৈনন্দিন বাংলা ব্যবহার করুন।
        আপনি প্রমিত বাংলা, আঞ্চলিক বাংলা অথবা কথ্য বাংলা ব্যবহার করতে পারেন।

        ⚠️ অনুগ্রহ করে আপনার নাম, ফোন নম্বর, ইমেইল, ঠিকানা বা অন্য কোনো ব্যক্তিগত
        পরিচয়মূলক তথ্য লিখবেন না।
        """
    )


    with st.form("participant_form"):

        age = st.selectbox(

            "Age group",

            [
                "Under 18",
                "18–24",
                "25–34",
                "35–44",
                "45 or above",
                "Prefer not to say",
            ],

        )


        education = st.selectbox(

            "Highest education",

            [
                "Secondary",
                "Higher Secondary",
                "Undergraduate",
                "Master's",
                "PhD",
                "Other",
                "Prefer not to say",
            ],

        )


        usage = st.selectbox(

            "How frequently do you use Bangla in everyday communication?",

            [
                "Almost always",
                "Very frequently",
                "Frequently",
                "Sometimes",
                "Rarely",
            ],

        )


        variety = st.selectbox(

            "Bangla variety primarily used",

            [
                "Standard Bangla",
                "Regional/Dialectal Bangla",
                "Mixture of Standard and Regional Bangla",
                "Other / Prefer not to say",
            ],

        )


        consent = st.checkbox(

            """
            I voluntarily agree to participate and allow my anonymous responses
            to be used for academic research.
            """

        )


        submitted = st.form_submit_button(
            "Start Questionnaire"
        )


        if submitted:

            if not consent:

                st.error(
                    "Please provide consent before starting."
                )

            else:

                st.session_state.participant = {

                    # Anonymous Participant ID
                    "participant_id": str(uuid.uuid4()),

                    "age": age,
                    "education": education,
                    "usage": usage,
                    "variety": variety,

                    # Overall questionnaire start time
                    "start_time": datetime.now(),

                }

                # Start first scenario timer
                st.session_state.scenario_start_time = datetime.now()

                st.session_state.started = True

                st.rerun()


# ==============================================================================
# SCENARIO QUESTIONNAIRE
# ==============================================================================

elif (
    st.session_state.index < len(SCENARIOS)
    and not st.session_state.usability_completed
):

    i = st.session_state.index

    category, context = SCENARIOS[i]


    # --------------------------------------------------------------------------
    # Progress Bar
    # --------------------------------------------------------------------------

    progress_value = (i + 1) / len(SCENARIOS)

    st.progress(progress_value)

    st.subheader(
        f"Scenario {i + 1} of {len(SCENARIOS)}"
    )


    # --------------------------------------------------------------------------
    # Break Reminder
    # --------------------------------------------------------------------------

    if i == 25:

        st.info(
            """
            ☕ আপনি প্রশ্নমালার অর্ধেক সম্পন্ন করেছেন!

            চাইলে কিছুক্ষণ বিরতি নিতে পারেন। তারপর Continue করে প্রশ্নমালা
            সম্পন্ন করুন।
            """
        )


    # --------------------------------------------------------------------------
    # Scenario Context
    # --------------------------------------------------------------------------

    st.info(
        f"**পরিস্থিতি (Situation):** {context}"
    )


    # --------------------------------------------------------------------------
    # Example
    # --------------------------------------------------------------------------

    with st.expander(
        "💡 কীভাবে উত্তর দেবেন? একটি উদাহরণ দেখুন"
    ):

        st.markdown(
            """
            **নমুনা পরিস্থিতি:**

            আপনি বন্ধুর সাথে ঘরে বসে আছেন। বাইরে খুব ঠান্ডা এবং জানালা খোলা।

            **১. আপনি যা বলবেন (Utterance):**

            *"আজকে আবহাওয়াটা একটু বেশিই ঠান্ডা, না?"*

            **২. আপনার আসল উদ্দেশ্য (Intended Meaning):**

            *"পরোক্ষভাবে বন্ধুকে জানালাটা বন্ধ করতে বলা।"*
            """
        )


    # --------------------------------------------------------------------------
    # Scenario Form
    # --------------------------------------------------------------------------

    with st.form(f"scenario_form_{i}"):


        # ----------------------------------------------------------------------
        # Question 1
        # ----------------------------------------------------------------------

        utterance = st.text_area(

            """
            ১. এই পরিস্থিতিতে আপনি স্বাভাবিকভাবে কী বলবেন?
            (What would you naturally say?)
            """,

            value=st.session_state.answers.get(
                (i, "u"),
                ""
            ),

            height=110,

            placeholder=(
                "উদাহরণ: আপনি বাস্তবে যেভাবে বলতেন সেভাবে লিখুন "
                "(আঞ্চলিক বা প্রমিত বাংলায়)..."
            ),

        )


        # ----------------------------------------------------------------------
        # Question 2
        # ----------------------------------------------------------------------

        intended = st.text_area(

            """
            ২. এই কথা বলে আপনি মূল কী বোঝাতে চেয়েছেন?
            (What was your intended meaning?)
            """,

            value=st.session_state.answers.get(
                (i, "m"),
                ""
            ),

            height=110,

            placeholder=(
                "উদাহরণ: আপনার কথার মাধ্যমে আসল কী উদ্দেশ্য, "
                "অনুভূতি বা অর্থ প্রকাশ করতে চেয়েছেন তা লিখুন..."
            ),

        )


        # ----------------------------------------------------------------------
        # Intent Categories
        # ----------------------------------------------------------------------

        intent_options = [

            "Not Specified",

            "অনুরোধ (Request)",

            "অস্বীকৃতি/না বলা (Refusal)",

            "ব্যঙ্গ/ঠাট্টা (Sarcasm/Irony)",

            "অভিযোগ (Complaint)",

            "প্রশংসা/ধন্যবাদ (Praise/Appreciation)",

            "সন্দেহ (Doubt)",

            "সহানুভূতি (Sympathy)",

            "হতাশা (Frustration)",

            "আশ্চর্য (Surprise)",

            "দুঃখ/নিরাশা (Disappointment)",

            "অন্যান্য (Other)",

        ]


        intent_cat = st.selectbox(

            """
            ৩. (ঐচ্ছিক) আপনার উদ্দেশ্যটি মূলত কোন ধরণের?
            (Optional: Select primary communicative function)
            """,

            intent_options,

            index=st.session_state.answers.get(
                (i, "cat_idx"),
                0
            ),

        )


        # ----------------------------------------------------------------------
        # Navigation Buttons
        # ----------------------------------------------------------------------

        c1, c2 = st.columns(2)

        back = c1.form_submit_button(
            "← Previous"
        )

        next_btn = c2.form_submit_button(
            "Save & Next →"
        )


        # ----------------------------------------------------------------------
        # Previous Button
        # ----------------------------------------------------------------------

        if back:

            st.session_state.answers[(i, "u")] = utterance

            st.session_state.answers[(i, "m")] = intended

            st.session_state.answers[(i, "cat_idx")] = (
                intent_options.index(intent_cat)
            )

            st.session_state.index = max(
                0,
                i - 1
            )

            # Reset scenario timer
            st.session_state.scenario_start_time = datetime.now()

            st.rerun()


        # ----------------------------------------------------------------------
        # Save & Next Button
        # ----------------------------------------------------------------------

        if next_btn:

            if not utterance.strip() or not intended.strip():

                st.error(
                    """
                    অনুগ্রহ করে দুটি প্রশ্নেরই উত্তর দিন।
                    (Please answer both text questions.)
                    """
                )

            else:

                # Save answers temporarily
                st.session_state.answers[(i, "u")] = utterance

                st.session_state.answers[(i, "m")] = intended

                st.session_state.answers[(i, "cat_idx")] = (
                    intent_options.index(intent_cat)
                )


                # --------------------------------------------------------------
                # Calculate Scenario Response Time
                # --------------------------------------------------------------

                if st.session_state.scenario_start_time:

                    response_time = (

                        datetime.now()
                        - st.session_state.scenario_start_time

                    ).total_seconds()

                else:

                    response_time = 0


                # --------------------------------------------------------------
                # Save Response Only Once
                # --------------------------------------------------------------

                if not st.session_state.answers.get(
                    (i, "saved"),
                    False
                ):

                    saved = save_response(

                        participant=st.session_state.participant,

                        scenario_id=i + 1,

                        category=category,

                        context=context,

                        utterance=utterance,

                        intended=intended,

                        intent_cat=intent_cat,

                        response_time=response_time,

                    )

                    if not saved:
                        st.session_state.started = False
                        st.session_state.participant = None
                        st.session_state.index = 0
                        st.session_state.answers = {}
                        st.session_state.scenario_start_time = None
                        st.session_state.usability_completed = False
                        st.rerun()

                    st.session_state.answers[(i, "saved")] = True


                # --------------------------------------------------------------
                # Move to Next Scenario
                # --------------------------------------------------------------

                st.session_state.index = i + 1

                # Start timer for next scenario
                st.session_state.scenario_start_time = datetime.now()

                st.rerun()


# ==============================================================================
# HCI / USABILITY EVALUATION
# ==============================================================================

elif not st.session_state.usability_completed:

    st.success(
        "আপনি মূল প্রশ্নমালা সম্পন্ন করেছেন!"
    )

    st.markdown(
        """
        # 📝 System Usability Feedback

        এখন আমাদের ডেটা সংগ্রহ সিস্টেমের ব্যবহারযোগ্যতা সম্পর্কে আপনার মতামত দিন।

        নিচের প্রতিটি বিবৃতির জন্য আপনার মতামত নির্বাচন করুন।

        **স্কেল:**

        - 1 = সম্পূর্ণ অসম্মত
        - 2 = অসম্মত
        - 3 = নিরপেক্ষ
        - 4 = সম্মত
        - 5 = সম্পূর্ণ সম্মত
        """
    )


    # --------------------------------------------------------------------------
    # Likert Scale
    # --------------------------------------------------------------------------

    scale = {

        "1 - সম্পূর্ণ অসম্মত": 1,

        "2 - অসম্মত": 2,

        "3 - নিরপেক্ষ": 3,

        "4 - সম্মত": 4,

        "5 - সম্পূর্ণ সম্মত": 5,

    }


    scale_options = list(scale.keys())


    with st.form("usability_form"):


        u1 = st.radio(
            "১. নির্দেশনাগুলো সহজে বুঝতে পেরেছি।",
            scale_options
        )


        u2 = st.radio(
            "২. পরিস্থিতিগুলো সহজে বুঝতে পেরেছি।",
            scale_options
        )


        u3 = st.radio(
            "৩. সিস্টেমটি ব্যবহার করা সহজ ছিল।",
            scale_options
        )


        u4 = st.radio(
            "৪. উদাহরণটি কীভাবে উত্তর দিতে হবে তা বুঝতে সাহায্য করেছে।",
            scale_options
        )


        u5 = st.radio(
            "৫. অগ্রগতির সূচকটি সহায়ক ছিল।",
            scale_options
        )


        u6 = st.radio(
            "৬. আমার প্রকৃত উদ্দেশ্য বা অর্থ প্রকাশ করা সহজ ছিল।",
            scale_options
        )


        u7 = st.radio(
            "৭. প্রশ্নমালার দৈর্ঘ্য যুক্তিসঙ্গত মনে হয়েছে।",
            scale_options
        )


        u8 = st.radio(
            "৮. সিস্টেমটি ব্যবহার করতে স্বাচ্ছন্দ্যবোধ করেছি।",
            scale_options
        )


        u9 = st.radio(
            "৯. ভবিষ্যতে অনুরূপ একটি সিস্টেম ব্যবহার করতে আগ্রহী।",
            scale_options
        )


        u10 = st.radio(
            "১০. সামগ্রিকভাবে আমি সিস্টেমটি নিয়ে সন্তুষ্ট।",
            scale_options
        )


        feedback = st.text_area(

            """
            এই সিস্টেমটি ব্যবহার করার অভিজ্ঞতা সম্পর্কে আপনার কোনো
            পরামর্শ বা মন্তব্য থাকলে লিখুন (ঐচ্ছিক):
            """,

            height=120,

        )


        submit_feedback = st.form_submit_button(
            "Submit Feedback"
        )


        # ----------------------------------------------------------------------
        # Save Usability Feedback
        # ----------------------------------------------------------------------

        if submit_feedback:

            participant = st.session_state.participant
            participant_id = participant_id_from_state(participant)

            if participant_id is None:
                st.error(
                    "Participant information is missing. Please restart the questionnaire and submit again."
                )
                st.stop()

            # --------------------------------------------------------------
            # Calculate Overall Completion Time
            # --------------------------------------------------------------

            completion_time = (

                datetime.now()
                - participant["start_time"]

            ).total_seconds() / 60


            # --------------------------------------------------------------
            # Save Usability Feedback
            # --------------------------------------------------------------

            save_usability_feedback(

                participant_id=participant_id,

                completion_time=completion_time,

                u1=scale[u1],
                u2=scale[u2],
                u3=scale[u3],
                u4=scale[u4],
                u5=scale[u5],
                u6=scale[u6],
                u7=scale[u7],
                u8=scale[u8],
                u9=scale[u9],
                u10=scale[u10],

                feedback=feedback,

            )


            st.session_state.usability_completed = True

            st.rerun()


# ==============================================================================
# FINAL THANK YOU PAGE
# ==============================================================================

else:

    st.success(
        """
        Thank you! Your responses and usability feedback
        have been successfully recorded.
        """
    )

    st.balloons()

    st.markdown(
        """
        ## আপনার অংশগ্রহণের জন্য অসংখ্য ধন্যবাদ। 🙏

        আপনার উত্তরগুলো বাংলা প্র্যাগম্যাটিক্স এবং ভাষাগত উদ্দেশ্য
        বিশ্লেষণ সম্পর্কিত গবেষণায় ব্যবহার করা হবে।
        """
    )


    # --------------------------------------------------------------------------
    # Start New Participant Response
    # --------------------------------------------------------------------------

    if st.button(
        "Start a New Response"
    ):

        st.session_state.started = False

        st.session_state.participant = None

        st.session_state.index = 0

        st.session_state.answers = {}

        st.session_state.scenario_start_time = None

        st.session_state.usability_completed = False

        st.rerun()

